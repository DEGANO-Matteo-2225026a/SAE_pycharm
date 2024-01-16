import pandas as pd
import openpyxl as op

"""
Le premier Intervenant renseigné dans chaque ressources est considéré par convention comme le responsable de celle-ci.

TableauDonnées = [[S1,Nom Ressource,Responsable True/False,Titulaire ,Acronyme, Titulaire True/False, NombreGroupes, TDNonD, TPD, Test],
                  [S2][R1.02][Responsable Oui/Non][Nom + Prenom][Acronyme][Titulaire Oui/Non][NombreGroupes][TDNonD][TPD][Test],
                  [S2][R2.01][Responsable Oui/Non][Nom + Prenom][Acronyme][Titulaire Oui/Non][NombreGroupes][TDNonD][TPD][Test]]
"""

# On charge le classeur Excel
QuiFaitQuoi = pd.ExcelFile('../Documents/QuiFaitQuoiPOURTEST.xlsx')
QuiFaitQuoiInfo = op.load_workbook('../Documents/QuiFaitQuoiPOURTEST.xlsx',data_only=True)

# On établie la liste des feuilles qui nous intéressent
ListeFeuillesQuiFaitQuoi = sorted(QuiFaitQuoi.sheet_names)

print(ListeFeuillesQuiFaitQuoi)

def RecuperationProfMatiere(FeuilleActuelle,TableauDonnées):

    # Active la page pour openPYXL
    Feuille = QuiFaitQuoiInfo[FeuilleActuelle]

    # Indice maximum Lignes et Colonnes
    CompteurLigne = Feuille.max_row + 1
    CompteurColonne = Feuille.max_column

    # Index pour le "curseur"
    IndexColonne = 2

    """
    print(CompteurColonne,CompteurLigne) # 11 Co 44 Li atm 
    print(Feuille.cell(1,1).internal_value)
    print(Feuille.cell(1,1).fill.start_color.index
    """

    for IndexLigne in range(2,CompteurLigne):

        # Permet de savoir lors du changement de couleur qu'une nouvelle matière apporte un nouveau responsable de module
        if (Feuille.cell(IndexLigne, IndexColonne).fill.start_color.index) != '00000000':
            MatiereActuelle = Feuille.cell(IndexLigne, IndexColonne-1).internal_value
            """
            print("NOUVELLE MATIERE : ", MatiereActuelle)
            """
            AlerteProf = True


        # Si nouvelle ligne de matière on ignore, si absence d'intervenant on ignore
        if (Feuille.cell(IndexLigne, IndexColonne).internal_value == None):
            continue

        # Récupère les données
        Intervenant = Feuille.cell(IndexLigne, IndexColonne).internal_value
        Acronyme = Feuille.cell(IndexLigne, IndexColonne+1).internal_value
        if Feuille.cell(IndexLigne, IndexColonne+1).internal_value == None:
            Acronyme = Intervenant
        if Feuille.cell(IndexLigne, IndexColonne+2).internal_value == "Oui":
            Titulaire = True
        else:
            Titulaire = False
        NombreGroupes = Feuille.cell(IndexLigne, IndexColonne+3).internal_value
        if Feuille.cell(IndexLigne, IndexColonne+3).internal_value == None:
            NombreGroupes = 0
        CM = Feuille.cell(IndexLigne, IndexColonne+4).internal_value
        if Feuille.cell(IndexLigne, IndexColonne+4).internal_value == None:
            CM = 0
        TDNonD = Feuille.cell(IndexLigne, IndexColonne+5).internal_value
        if Feuille.cell(IndexLigne, IndexColonne+5).internal_value == None:
            TDNonD = 0
        TPD = Feuille.cell(IndexLigne, IndexColonne+6).internal_value
        if Feuille.cell(IndexLigne, IndexColonne+6).internal_value == None:
            TPD = 0
        Test = Feuille.cell(IndexLigne, IndexColonne+7).internal_value
        if Feuille.cell(IndexLigne, IndexColonne+7).internal_value == None:
            Test = 0

        """
        if AlerteProf is not False:
            print(Intervenant,"est responsable de", MatiereActuelle)
        """

        # On remplis le tableau avec les données sur notre nouveau prof
        TableauDonnées.append([Feuille.title,MatiereActuelle,AlerteProf,Intervenant,Acronyme,Titulaire,NombreGroupes,CM,TDNonD,TPD,Test])

        # On se rappel de remettre l'alerte à 0
        AlerteProf = False

        """
        print("Matière = ", MatiereActuelle,"," ,"Intervenant = ",Intervenant,"," ," Acronyme = ",Acronyme,"," ,
              "Titulaire = ",Titulaire,",","NombreGroupes = ",NombreGroupes,",",
              "TDNonD = ",TDNonD,",","TPD = ",TPD,",","Test = ",Test,)
        """

# On parcours toutes les feuilles
def RecuperationParFeuille(ListeFeuilles):

    # Tableau où seront rangées les données
    TableauDonnées = []

    # On parcours toutes les feuilles fournis
    for Feuille in ListeFeuilles:

        RecuperationProfMatiere(Feuille,TableauDonnées)

        """
        print("")
        print("PAGE SUIVANTE")
        print("PAGE SUIVANTE")
        print("PAGE SUIVANTE")
        print("")
        """

    return TableauDonnées

DonneesQuiFaitQuoi = RecuperationParFeuille(ListeFeuillesQuiFaitQuoi)
print(DonneesQuiFaitQuoi)