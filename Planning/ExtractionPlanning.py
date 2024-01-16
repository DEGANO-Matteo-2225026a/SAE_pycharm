import openpyxl as op
from datetime import datetime

"""

TableauDonnees[0] = [[['FFCC99FF', 'R1.01', 18, 30, 38, 'Acronyme'],
                   ['FFAECF00', 'R1.06-1', 8, 11, 3, 'Acronyme'],
                   ['FF66CCFF', 'R1.02', 6, 2, 16, 'Acronyme'],
                   ['FFAECF00', 'R1.06-2', 4, 6, 4, 'Acronyme'],
                   ['FFFFF200', 'R1.03', 4, 8, 8, 'Acronyme']],
                  [['FFCC99FF', 'R3.01', 8, 0, 18, 'Acronyme'],
                   ['FFAECF00', 'R3.08', 10, 14, 6, 'Acronyme']]]
                   
TableauDonnees[1] = [[Semaine, Ressource, TypeCours, TypeSalle],
                  [Semaine, Ressource, TypeCours, TypeSalle],
                  [[Semaine, Ressource, TypeCours, TypeSalle],
                  [Semaine, Ressource, TypeCours, TypeSalle],
                  [[Semaine, Ressource, TypeCours, TypeSalle]]

Les données du document planning nous apparaissent avec cette logique :
- Les données concernant les matières de chaque pages de semestre sont divisées en deux parties, la deuxième commençant 
après la colonne contenant "A placer dans EDT"
- Ces deux parties servent uniquement à faciliter le repèrage des matières, elles ont bien lieux dans la même semaine.
- On remarque un compteur d'heure en Amphi, Salle Machine (Y) et Test. Ce compteur nous permet de trouver (sauf erreurs)
le total d'heure en Salle de TD -> Heures Totales - ((Amphi + Y) * 2) 

On possède par les info suivantes :
- Nom du semestre
- Date Semaine
- Numéro de Semaine
- Cours
- TD
- TP
- Test
- Sae
- Heures Info
- Heures Non Info
- Heures Totales

ROUGE = SEMAINE MORTE DONC OSEF
"""

# On charge le classeur Excel
# import pandas as pd
# Planning = pd.ExcelFile('../Documents/Planning_2023-2024.xlsx')
PlanningInfo = op.load_workbook('../Documents/Planning_2023-2024-2.xlsx',data_only=True)



def PurgeFeuille(Planning):
    FeuilleASupprimer = []

    for Feuille in Planning.sheetnames:
        if Feuille[0] != "S" or not Feuille[1].isdigit():
            FeuilleASupprimer.append(Feuille)

    for Feuille in FeuilleASupprimer:
        Planning.remove(Planning[Feuille])



def LocateDate(Feuille):

    # On recherche l'Indice de la colonne des Dates
    IndiceDate = 1
    while Feuille.cell(1,IndiceDate).value != "Date":
        IndiceDate += 1

    # On cherche le nombre de semaine pour nos boucles.
    LongueurDate = 1
    BlancsDate = 0
    while BlancsDate != 3:
        if Feuille.cell(LongueurDate, IndiceDate).value != None:
            BlancsDate = 0
        else:
            BlancsDate += 1
        LongueurDate += 1

    # On oublie pas de retirer les lignes vides servant à confirmer la fin des infos de la colonne
    LongueurDate = LongueurDate - 4

    return IndiceDate, LongueurDate



def LocateLimite(Feuille,ColonneDate):

    LimiteGauche = 1
    LimiteDroite = ColonneDate

    # On cherche l'indice du début de la première colonne "Cours"
    while Feuille.cell(1,LimiteGauche).value != "Cours":
        LimiteGauche += 1

    # On cherche l'indice de fin de la deuxième colonne "Test"
    while Feuille.cell(1,LimiteDroite-1).value != "Test" or Feuille.cell(1,LimiteDroite).fill.start_color.index != '00000000':
        LimiteDroite += 1

    return LimiteDroite,LimiteGauche



def GetRessources(Feuille,LimiteBoucle,LimiteDroite):

    IndiceLigne = LimiteBoucle + 2
    IndiceColonne = LimiteDroite
    IndiceTableau = 0

    TableauRessource = []

    for i in range(1,LimiteBoucle):
        for j in range(1,LimiteDroite):
            if Feuille.cell(IndiceLigne + i,j).fill.start_color.index != '00000000' and Feuille.cell(IndiceLigne + i,j).value == 'X':

                TableauRessource.append([])

                # On ajoute la couleur de la Ressource au tableau nécessaire pour GetInfoPlanning()
                TableauRessource[IndiceTableau].append(Feuille.cell(IndiceLigne + i,j).fill.start_color.index)

                # Indice du nom de la matière d'après ce qui est écrit
                TrouveNom = 1
                while Feuille.cell(IndiceLigne + i, j + TrouveNom).value == None:
                    TrouveNom += 1

                # On ajoute le nom de la Ressource au tableau
                TableauRessource[IndiceTableau].append(Feuille.cell(IndiceLigne + i, j + TrouveNom).value)

                # Nombre de CM de la matière d'après ce qui est écrit
                ValeurCM = 0
                if Feuille.cell(IndiceLigne + i, j + TrouveNom + 3).value != None:
                    ValeurCM = Feuille.cell(IndiceLigne + i, j + TrouveNom + 3).value

                # On ajoute ValeurCM au tableau
                TableauRessource[IndiceTableau].append(ValeurCM)

                # Nombre de TD de la matière d'après ce qui est écrit
                ValeurTD = 0
                if Feuille.cell(IndiceLigne + i, j + TrouveNom + 5).value != None:
                    ValeurTD = Feuille.cell(IndiceLigne + i, j + TrouveNom + 5).value

                # On ajoute ValeurTD au tableau
                TableauRessource[IndiceTableau].append(ValeurTD)

                # Nombre de TP de la matière d'après ce qui est écrit
                ValeurTP = 0
                if Feuille.cell(IndiceLigne + i, j + TrouveNom + 7).value != None:
                    ValeurTP = Feuille.cell(IndiceLigne + i, j + TrouveNom + 7).value

                # On ajoute ValeurTP au tableau
                TableauRessource[IndiceTableau].append(ValeurTP)

                # Acronyme du Responsable de la matière d'après ce qui est écrit
                Responsable = "Personne"
                if Feuille.cell(IndiceLigne + i, j + TrouveNom + 10).value != None:
                    Responsable = Feuille.cell(IndiceLigne + i, j + TrouveNom + 10).value

                # On ajoute l'Acronyme du Responsable au tableau
                TableauRessource[IndiceTableau].append(Responsable)

                # On oublie pas de changer l'indice du tableau pour passer à la ressource suivante
                IndiceTableau += 1

    return TableauRessource



def GetInfoPlanning(Feuille,ColonneDate,LimiteGauche,LimiteDroite,LimiteBoucle,TableauRessource,IndicePage,TableauInfoPlanning):

    # On parcours toutes les lignes de la colonne Date
    for i in range(2,LimiteBoucle+1):
        if Feuille.cell(i,ColonneDate).fill.start_color.index == 'FFFF0000' or Feuille.cell(i,ColonneDate).value == None:
            continue

        # On ignore les couleurs qui ne sont pas dans notre liste de matière
        for k in range (len(TableauRessource[IndicePage])):
            if Feuille.cell(i,ColonneDate).fill.start_color.index not in TableauRessource[k][0]:
                continue

        SemaineActuelle = Feuille.cell(i, ColonneDate).value.strftime('%d-%m-%Y')

        # On parcours toutes les cellules de la ligne
        for j in range(LimiteGauche,LimiteDroite + 1):

            if Feuille.cell(i,j).value not in {"X","Y"} or Feuille.cell(i,j).fill.start_color.index == 'FFCCCCCC':
                continue

            TypeCours = Feuille.cell(1,j).value
            IndiceCelluleCours = 0

            while TypeCours == None:
                if Feuille.cell(1,j-IndiceCelluleCours).value == None:
                    IndiceCelluleCours += 1
                TypeCours = Feuille.cell(1,j-IndiceCelluleCours).value

            Couleur = Feuille.cell(i, j).fill.start_color.index

            for k in range(len(TableauRessource)):
                if Feuille.cell(i, j).fill.start_color.index == TableauRessource[k][0]:
                    Couleur = TableauRessource[k][1]

            TableauInfoPlanning.append([SemaineActuelle,Couleur,TypeCours,Feuille.cell(i,j).value])
            # print(SemaineActuelle,Couleur,TypeCours,Feuille.cell(i,j).value)
    return TableauInfoPlanning

def RecuperationDonneesFeuille(TableauDonnees,TableauInfoPlanning,FeuilleActuelle, IndicePage):

    # Active la page pour openPYXL
    Feuille = PlanningInfo[FeuilleActuelle.title]

    # print("VALEUR GRIS",Feuille.cell(2,2).fill.start_color.index)

    # On récupère les informations de la colonne Date et de la longueur du tableau
    ColonneDate = LocateDate(Feuille)
    LimiteBoucle = ColonneDate[1]
    ColonneDate = ColonneDate[0]

    # On récupère les limites pour les informations du tableau
    LimiteDroite = LocateLimite(Feuille,ColonneDate)
    LimiteGauche = LimiteDroite[1]
    LimiteDroite = LimiteDroite[0]

    TableauRessource = GetRessources(Feuille, LimiteBoucle, LimiteDroite)
    TableauDonnees.append(TableauRessource)

    GetInfoPlanning(Feuille, ColonneDate, LimiteGauche, LimiteDroite, LimiteBoucle, TableauRessource, IndicePage,TableauInfoPlanning)

    IndicePage += 1

    return



# Automatise le changement de feuilles
def RecuperationParFeuille(ListeFeuilles):
    TableauDonnees = []
    TableauInfoPlanning = []
    IndicePage = 0
    for Feuille in ListeFeuilles:
        RecuperationDonneesFeuille(TableauDonnees,TableauInfoPlanning,Feuille,IndicePage)
    return TableauDonnees, TableauInfoPlanning



# On établie la liste des feuilles qui nous intéressent
PurgeFeuille(PlanningInfo)

# On Lance le code
TableauDonnees = RecuperationParFeuille(PlanningInfo)
print(TableauDonnees[0])
print(TableauDonnees[1])

