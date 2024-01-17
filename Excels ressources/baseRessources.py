import openpyxl
from openpyxl.styles import Font

# Créer un nouveau classeur Excel
classeur = openpyxl.Workbook()

# Sélectionner la feuille active
feuille = classeur.active

# Ajouter des données à la feuille
feuille['A1'] = 'Ressource'
feuille['A1'].font = Font(bold=True)

feuille['F1'] = 'Responsable'
feuille['F1'].font = Font(bold=True)

feuille['A3'] = 'Maquette'
feuille['B3'] = 'CM (h)'
feuille['C3'] = 'TD (h)'
feuille['D3'] = 'TP (h)'

feuille['A6'] = 'Intervenants'
feuille['A6'].font = Font(bold=True)

feuille['A8'] = 'Repartition'
feuille['A8'].font = Font(bold=True)
feuille['B8'] = 'Nombre de Groupes'
feuille['A9'] = 'CM'
feuille['A11'] = 'TD'
feuille['A13'] = 'TP dedoubles'
feuille['A15'] = 'TP non dedoubles'

feuille['A17'] = 'Organisation detaillee'
feuille['A17'].font = Font(bold=True)
feuille['B19'] = 'CM heures'
feuille['C19'] = 'TD heures'
feuille['D19'] = 'TP heures'
feuille['E19'] = 'Test'

feuille['A21'] = 'Service prévisionnel vacataires'
feuille['A22'] = 'Nom'
feuille['C22'] = 'BUT1/BUT2/BUT3'
feuille['D22'] = 'ParcoursA  ParcoursB'
feuille['E22'] = 'FI  FA'
feuille['F22'] = 'Nombres d\'heures prévuesde septembre à décembre'
feuille['L22'] = 'Nombres d\'heures prévuesde janvier à août'
feuille['F23'] = 'CM'
feuille['G23'] = 'TD'
feuille['H23'] = 'TP(en 1/2 groupe)'
feuille['I23'] = 'TP(nondédoublé)'
feuille['J23'] = 'TP àdéclarerARES'
feuille['K23'] = 'Total enHETD'
feuille['L23'] = 'CM'
feuille['M23'] = 'TD'
feuille['N23'] = 'TP(en 1/2 groupe)'
feuille['O23'] = 'TP(nondédoublé)'
feuille['P23'] = 'TP àdéclarerARES'
feuille['Q23'] = 'Total enHETD'
feuille['R23'] = 'Total enHETD'

feuille.merge_cells('A22:B23')
feuille.merge_cells('C22:C23')
feuille.merge_cells('D22:D23')
feuille.merge_cells('E22:E23')
feuille.merge_cells('F22:K22')
feuille.merge_cells('L22:Q22')

feuille['A25'] = 'Service prévisionnel titulaires'
feuille['A26'] = 'Nom'
feuille['C26'] = 'BUT1/BUT2/BUT3'
feuille['D26'] = 'ParcoursA  ParcoursB'
feuille['E26'] = 'FI  FA'
feuille['F26'] = "Nombres d'heures prévues de septembre à décembre"
feuille['L26'] = "Nombres d'heures prévues de janvier à août"
feuille['F27'] = 'CM'
feuille['G27'] = 'TD'
feuille['H27'] = 'TP(en 1/2 groupe)'
feuille['I27'] = 'TP(nondédoublé)'
feuille['J27'] = 'TP à déclarer ARES'
feuille['K27'] = 'Total en HETD'
feuille['L27'] = 'CM'
feuille['M27'] = 'TD'
feuille['N27'] = 'TP(en 1/2 groupe)'
feuille['O27'] = 'TP(nondédoublé)'
feuille['P27'] = 'TP à déclarer ARES'
feuille['Q27'] = 'Total en HETD'
feuille['R27'] = 'Total en HETD'

feuille.merge_cells('A26:B27')
feuille.merge_cells('C26:C27')
feuille.merge_cells('D26:D27')
feuille.merge_cells('E26:E27')
feuille.merge_cells('F26:K26')
feuille.merge_cells('L26:Q26')


# Enregistrer le classeur dans un fichier
classeur.save('Ressources.xlsx')
