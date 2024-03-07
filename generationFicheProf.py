class GenerationFicheProf:
    def run(self):
        import openpyxl
        from openpyxl import load_workbook
        from pathlib import Path
        from openpyxl.styles import Font

        # Creer un nouveau classeur Excel
        classeur = load_workbook("./Excels ressources/FicheProf.xlsx")

        # Selectionner la feuille active
        ancienne_feuille = classeur.active

        recherche = load_workbook("./Excels ressources/FichierRessources.xlsx")
        f = recherche.active



        def ficheProf(prof):

            # Supprimer les feuilles existantes
            for i in classeur.worksheets:
                if i.title == prof:
                    classeur.remove(classeur[prof])

            # Créer une nouvelle feuille pour chaque professeur
            feuille = classeur.copy_worksheet(ancienne_feuille)
            feuille.title = prof

            # nom professeur
            cellule_selectionnee = feuille['B1']
            cellule_selectionnee.value = prof

            
            for f in recherche:
                nomprof = trouver_ligne(f, prof, 1)
                if nomprof is None:
                    continue
                # On cherche si le professeur est responsable de la matière
                respounon = f.cell(row=nomprof - 1, column=1).value

                # Si oui on la place responsable
                if respounon == "Intervenants":
                    feuille.insert_rows(4)
                    feuille.cell(row=4, column=1, value=f.title)
                # Sinon on la place intervenant
                else:
                    rowInter = trouver_ligne(feuille, "Interventions", 1)
                    feuille.insert_rows(rowInter + 1)
                    feuille.cell(row=rowInter + 1, column=1, value=f.title)

                # On récupère les volumes horaires pour chaque ressource
                horaires = trouver_ligne(f, prof, 11)
                if horaires is None:
                    continue
                ligneH = trouver_ligne(feuille, "Matieres", 1)
                feuille.insert_rows(ligneH + 1)

                cm1 = f.cell(row=horaires, column=6).value
                if cm1 is None:
                    cm1 = 0
                cm2 = f.cell(row=horaires, column=12).value
                if cm2 is None:
                    cm2 = 0
                td1 = f.cell(row=horaires, column=7).value
                if td1 is None :
                    td1 = 0
                td2 = f.cell(row=horaires, column=13).value
                if td2 is None:
                    td2 = 0
                tpd1 = f.cell(row=horaires, column=8).value
                if tpd1 is None:
                    tpd1 = 0
                tpd2 = f.cell(row=horaires, column=14).value
                if tpd2 is None:
                    tpd2 = 0
                tp1 = f.cell(row=horaires, column=9).value
                if tp1 is None or tp1 is not int:
                    tp1 = 0
                tp2 = f.cell(row=horaires, column=15).value
                if tp2 is None:
                    tp2 = 0
                ares1 = f.cell(row=horaires, column=10).value
                if ares1 is None:
                    ares1 = 0
                ares2 = f.cell(row=horaires, column=16).value
                if ares2 is None:
                    ares2 = 0
                total = f.cell(row=horaires, column=18).value
                if total is None:
                    total = 0


                feuille.cell(row=ligneH + 1, column=1, value=f.title)
                feuille.cell(row=ligneH + 1, column=3, value=cm1+cm2)
                feuille.cell(row=ligneH + 1, column=4, value=(td1+td2))
                feuille.cell(row=ligneH + 1, column=5, value=(tpd1+tpd2))
                feuille.cell(row=ligneH + 1, column=6, value=(tp1+tp2))
                feuille.cell(row=ligneH + 1, column=7, value=(ares1+ares2))
                feuille.cell(row=ligneH + 1, column=8, value=total)

                # On insère l'année
                cellule_selectionnee = f.cell(row=horaires, column=3).value
                feuille.cell(row=ligneH + 1, column=2, value=cellule_selectionnee)





        def trouver_ligne(feuille, contenu_cible, debut):
            for row_number, row in enumerate(feuille.iter_rows(values_only=True), start=1):
                if row_number >= debut and contenu_cible in row:
                    return row_number

            return None

        liste_prof = []
        for p in recherche:
            for i in range(7, p.max_row):
                prof = p.cell(row=i, column=1).value
                if prof is None:
                    break
                if prof in liste_prof:
                    continue
                else:
                    liste_prof.append(prof)
                    ficheProf(prof)


        classeur.save("./Excels ressources/FicheProf.xlsx")

def main():
    # Instanciation de la classe GenerationFicheProf
    prof_generator = GenerationFicheProf()
    # Appel de la méthode run()
    prof_generator.run()

if __name__ == "__main__":
    main()

