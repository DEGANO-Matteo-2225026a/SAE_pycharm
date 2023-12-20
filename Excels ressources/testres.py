import openpyxl
from openpyxl import load_workbook

# Charger le classeur Excel existant
classeur_existant = load_workbook('base.xlsx')

# Sélectionner la feuille existante (ou créer une nouvelle feuille si elle n'existe pas)
nom_feuille = 'Sheet'  # Remplacez par le nom de votre feuille
feuille = classeur_existant[nom_feuille] if nom_feuille in classeur_existant.sheetnames else classeur_existant.create_sheet(nom_feuille)

cellule_selectionnee = feuille['C1']
cellule_selectionnee.value = 'Nomres'

cellule_selectionnee = feuille['H1']
cellule_selectionnee.value = 'Nomresp'

cellule_selectionnee = feuille['B4']
cellule_selectionnee.value = '...h'

cellule_selectionnee = feuille['C4']
cellule_selectionnee.value = '...h'

cellule_selectionnee = feuille['D4']
cellule_selectionnee.value = '...h'

ligne_insertion = 7
feuille.insert_rows(ligne_insertion)
feuille.cell(row=ligne_insertion, column=1, value='prof1')

ligne_insertion = 8
feuille.insert_rows(ligne_insertion)
feuille.cell(row=ligne_insertion, column=1, value='prof2')

ligne_insertion = 9
feuille.insert_rows(ligne_insertion)
feuille.cell(row=ligne_insertion, column=1, value='prof3')

# à finir avec bdd

# Enregistrer le classeur modifié dans un nouveau fichier
classeur_existant.save('fichier_modifie.xlsx')
