class GenerationRessources:
    def run(self):
        import openpyxl
        from openpyxl import load_workbook
        from pathlib import Path
        from openpyxl.styles import Font
        import sqlite3

        # connection à la base de donnée
        db_connection = sqlite3.connect("SAE.db")
        cursor = db_connection.cursor()

        # Creer un nouveau classeur Excel
        classeur = load_workbook("./Excels ressources/FichierRessources.xlsx")

        # Selectionner la feuille active
        ancienne_feuille = classeur.active

        def Remplissage(res, cmtot, tdtot, tptot, resp):

            # Supprimer les feuilles existantes
            for i in classeur.worksheets:
                if i.title == res:
                    classeur.remove(classeur[res])

            # Créer une nouvelle feuille pour chaque matière
            feuille = classeur.copy_worksheet(ancienne_feuille)
            feuille.title = res

            # nom responsable de ressource
            cellule_selectionnee = feuille['H1']
            cellule_selectionnee.value = resp

            # nom ressource
            cellule_selectionnee = feuille['C1']
            cellule_selectionnee.value = res

            # volume horaire total
            cellule_selectionnee = feuille['B4']
            cellule_selectionnee.value = cmtot

            cellule_selectionnee = feuille['C4']
            cellule_selectionnee.value = tdtot

            cellule_selectionnee = feuille['D4']
            cellule_selectionnee.value = tptot

            # Récupération des donnees des intervenants
            requete = "SELECT Intervenant, Acronyme, CM, NombreGroupes FROM DONNEEPROF WHERE SUBSTR(MatiereActuelle, 1, INSTR(MatiereActuelle, ' ') - 1) = ?"
            cursor.execute(requete, (res,))
            liste_inter = cursor.fetchall()

            # insertion données dans la feuille
            for i in range(len(liste_inter)):
                # nom et acronyme
                inter = trouver_ligne(feuille, "Intervenants")

                ligne_insertion = 7 + i
                feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][0])
                feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][1])

                # pourcentage du total de cm effectué par le professeur
                cms = trouver_ligne(feuille, "CM ")

                ligne_insertion = cms + 1 + i
                feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][1])
                feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][2])

                tds = trouver_ligne(feuille, "TD ")

                # nombre de groupes en TD
                ligne_insertion = tds + 1 + i
                feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][1])
                feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][3])

                tpsD = trouver_ligne(feuille, "TP dedoubles")

                # nombre de groupes en TP dédoublés
                ligne_insertion = tpsD + 1 + i
                feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][1])
                feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][3])

                tps = trouver_ligne(feuille, "TP non  dedoubles")

                # nombre de groupes en TP non dédoublés
                ligne_insertion = tps + 1 + i
                feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][1])
                feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][3])

            # Récupération des données de cours par semaine
            cours = "SELECT Semaine, TypeCours, COUNT(*) AS NombreCours FROM PLANINFO WHERE Ressource = ? AND TypeCours IS NOT NULL GROUP BY Semaine, TypeCours "
            cursor.execute(cours, (res,))
            liste_cours = cursor.fetchall()

            dates = []
            for i in range(len(liste_cours)):

                hcm = trouver_ligne(feuille, "CM (h)")
                ligne_insertion = hcm + 1 + i
                feuille.insert_rows(ligne_insertion)
                feuille.cell(row=ligne_insertion, column=1, value=liste_cours[i][0])

                # feuille.insert_rows(ligne_insertion)
                # feuille.cell(row=ligne_insertion, column=1, value=liste_cours[i][0])

                ligne_date = trouver_ligne(feuille, liste_cours[i][0])
                if ligne_date is None:
                    continue

                if liste_cours[i][1] == "Cours":
                    feuille.cell(row=ligne_date, column=2, value=(liste_cours[i][2]) * 2)
                elif liste_cours[i][1] == "TD":
                    feuille.cell(row=ligne_date, column=3, value=(liste_cours[i][2] * 2))
                elif liste_cours[i][1] == "TP":
                    feuille.cell(row=ligne_date, column=4, value=(liste_cours[i][2] * 2))
                else:
                    feuille.cell(row=ligne_date, column=5, value=(liste_cours[i][2] * 2))


            # Récupération des données pour remplir les tableaux titulaire de septembre à décembre
            cours = "SELECT TypeCours, COUNT(*) FROM PLANINFO WHERE Ressource = ? AND (Semaine LIKE '%-09-%' OR Semaine LIKE '%-10-%' OR Semaine LIKE '%-11-%' OR Semaine LIKE '%-12-%') GROUP BY TypeCours"
            cursor.execute(cours, (res,))
            cours_s1 = cursor.fetchall()

            # Récupération des données pour remplir les tableaux vacataire de janvier à aout
            cours = "SELECT TypeCours, COUNT(*) FROM PLANINFO WHERE Ressource = ? AND (Semaine NOT LIKE '%-09-%' AND Semaine NOT LIKE '%-10-%' AND Semaine NOT LIKE '%-11-%' AND Semaine NOT LIKE '%-12-%') GROUP BY TypeCours"
            cursor.execute(cours, (res,))
            cours_s2 = cursor.fetchall()

            # Récupération des données du titulaire
            prof = "SELECT Feuille_title,  SUBSTR(MatiereActuelle, 1, INSTR(MatiereActuelle, ' ') - 1) AS matiere, Intervenant, CM, TD, TDNonD, TPD FROM DONNEEPROF WHERE matiere = ? AND AlerteProf = 1"
            cursor.execute(prof, (res,))
            liste_titu = cursor.fetchall()

            ligne_titu = trouver_ligne(feuille, "Service previsionnel titulaires")

            ligne_insertion = ligne_titu + 3

            if len(liste_titu) > 0:
                feuille.cell(row=ligne_insertion, column=1, value=(liste_titu[0][2]))
                if "S1" in liste_titu[0][0] or "S2" in liste_titu[0][0]:
                    feuille.cell(row=ligne_insertion, column=3, value="BUT1")
                elif "S3" in liste_titu[0][0] or "S4" in liste_titu[0][0]:
                    feuille.cell(row=ligne_insertion, column=3, value="BUT2")
                else:
                    feuille.cell(row=ligne_insertion, column=3, value="BUT3")
                if "A" in liste_titu[0][1]:
                    feuille.cell(row=ligne_insertion, column=4, value="A")
                elif "B" in liste_titu[0][1]:
                    feuille.cell(row=ligne_insertion, column=4, value="B")

                # On remplie les bonnes cases de septembre à décembre
                if len(cours_s1) > 0:
                    if cours_s1[0][0] == "Cours":
                        feuille.cell(row=ligne_insertion, column=6, value=(cours_s1[0][1] * 2 * liste_titu[0][3]))
                    elif cours_s1[0][0] == "TD":
                        feuille.cell(row=ligne_insertion, column=7, value=(cours_s1[0][1] * 2 * liste_titu[0][4]))
                    elif cours_s1[0][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=8, value=(cours_s1[0][1] * 2 * liste_titu[0][5]))
                        feuille.cell(row=ligne_insertion, column=9, value=(cours_s1[0][1] * 2 * liste_titu[0][6]))

                if len(cours_s1) > 1:
                    if cours_s1[1][0] == "TD":
                        feuille.cell(row=ligne_insertion, column=7, value=(cours_s1[1][1] * 2 * liste_titu[0][4]))
                    elif cours_s1[1][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=8, value=(cours_s1[1][1] * 2 * liste_titu[0][5]))
                        feuille.cell(row=ligne_insertion, column=9, value=(cours_s1[1][1] * 2 * liste_titu[0][6]))

                if len(cours_s1) > 2:
                    if cours_s1[2][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=8, value=(cours_s1[2][1] * 2 * liste_titu[0][5]))
                        feuille.cell(row=ligne_insertion, column=9, value=(cours_s1[2][1] * 2 * liste_titu[0][6]))

                # On remplie les bonnes cases de janvier à aout

                if len(cours_s2) > 0:
                    if cours_s2[0][0] == "Cours":
                        feuille.cell(row=ligne_insertion, column=12, value=(cours_s2[0][1] * 2 * liste_titu[0][3]))
                    elif cours_s2[0][0] == "TD":
                        feuille.cell(row=ligne_insertion, column=13, value=(cours_s2[0][1] * 2 * liste_titu[0][4]))
                    elif cours_s2[0][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=14, value=(cours_s2[0][1] * 2 * liste_titu[0][5]))
                        feuille.cell(row=ligne_insertion, column=15, value=(cours_s2[0][1] * 2 * liste_titu[0][6]))

                if len(cours_s2) > 1:
                    if cours_s2[1][0] == "TD":
                        feuille.cell(row=ligne_insertion, column=13, value=(cours_s2[1][1] * 2 * liste_titu[0][4]))
                    elif cours_s2[1][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=14, value=(cours_s2[1][1] * 2 * liste_titu[0][5]))
                        feuille.cell(row=ligne_insertion, column=15, value=(cours_s2[1][1] * 2 * liste_titu[0][6]))

                if len(cours_s2) > 2:
                    if cours_s2[2][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=14, value=(cours_s2[2][1] * 2 * liste_titu[0][5]))
                        feuille.cell(row=ligne_insertion, column=15, value=(cours_s2[2][1] * 2 * liste_titu[0][6]))

                valeur_cellule_6 = feuille.cell(row=ligne_insertion, column=6).value
                valeur_cellule_7 = feuille.cell(row=ligne_insertion, column=7).value
                valeur_cellule_8 = feuille.cell(row=ligne_insertion, column=8).value
                valeur_cellule_9 = feuille.cell(row=ligne_insertion, column=9).value

                # Vérification si les cellules ne sont pas vides
                if valeur_cellule_6 is None:
                    valeur_cellule_6 = 0
                if valeur_cellule_7 is None:
                    valeur_cellule_7 = 0
                if valeur_cellule_8 is None:
                    valeur_cellule_8 = 0
                if valeur_cellule_9 is None:
                    valeur_cellule_9 = 0

                # Calcul de hetd_ares_1
                hetd_ares_1 = valeur_cellule_8 + valeur_cellule_9

                # Assignation de la valeur à la cellule 10
                feuille.cell(row=ligne_insertion, column=10, value=hetd_ares_1)

                hetd_total_1 = valeur_cellule_6 * 1.5 + valeur_cellule_7 + valeur_cellule_8 + valeur_cellule_9
                feuille.cell(row=ligne_insertion, column=11, value=hetd_total_1)

                valeur_cellule_12 = feuille.cell(row=ligne_insertion, column=12).value
                valeur_cellule_13 = feuille.cell(row=ligne_insertion, column=13).value
                valeur_cellule_14 = feuille.cell(row=ligne_insertion, column=14).value
                valeur_cellule_15 = feuille.cell(row=ligne_insertion, column=15).value

                if valeur_cellule_12 is None:
                    valeur_cellule_12 = 0
                if valeur_cellule_13 is None:
                    valeur_cellule_13 = 0
                if valeur_cellule_14 is None:
                    valeur_cellule_14 = 0
                if valeur_cellule_15 is None:
                    valeur_cellule_15 = 0

                # Calcul de hetd_ares_1
                hetd_ares_2 = valeur_cellule_14 + valeur_cellule_15

                # Assignation de la valeur à la cellule 16
                feuille.cell(row=ligne_insertion, column=16, value=hetd_ares_2)

                hetd_total_1 = valeur_cellule_12 * 1.5 + valeur_cellule_13 + valeur_cellule_14 + valeur_cellule_15
                feuille.cell(row=ligne_insertion, column=17, value=hetd_total_1)

                valeur_cellule_11 = feuille.cell(row=ligne_insertion, column=11).value
                valeur_cellule_17 = feuille.cell(row=ligne_insertion, column=17).value
                if valeur_cellule_11 is None:
                    valeur_cellule_11 = 0
                if valeur_cellule_17 is None:
                    valeur_cellule_17 = 0

                hetd_final = valeur_cellule_11 + valeur_cellule_17
                feuille.cell(row=ligne_insertion, column=18, value=hetd_final)

            # Récupération des données des vacataires
            prof = "SELECT Feuille_title,  SUBSTR(MatiereActuelle, 1, INSTR(MatiereActuelle, ' ') - 1) AS matiere, Intervenant, CM, TD, TDNonD, TPD FROM DONNEEPROF WHERE matiere = ? AND AlerteProf = 0"
            cursor.execute(prof, (res,))
            liste_vac = cursor.fetchall()

            ligne_vac = trouver_ligne(feuille, "Service previsionnel vacataires")

            for i in range(len(liste_vac)):
                ligne_insertion = ligne_vac + i + 3
                feuille.cell(row=ligne_insertion, column=1, value=(liste_vac[i][2]))
                if "S1" in liste_vac[i][0] or "S2" in liste_vac[i][0]:
                    feuille.cell(row=ligne_insertion, column=3, value="BUT1")
                elif "S3" in liste_vac[i][0] or "S4" in liste_vac[i][0]:
                    feuille.cell(row=ligne_insertion, column=3, value="BUT2")
                else:
                    feuille.cell(row=ligne_insertion, column=3, value="BUT3")
                if "A" in liste_vac[i][1]:
                    feuille.cell(row=ligne_insertion, column=4, value="A")
                elif "B" in liste_vac[i][1]:
                    feuille.cell(row=ligne_insertion, column=4, value="B")

                # On remplie les bonnes cases de septembre à décembre
                if len(cours_s1) > 0:
                    if cours_s1[0][0] == "Cours":
                        feuille.cell(row=ligne_insertion, column=6, value=(cours_s1[0][1] * 2 * liste_vac[i][3]))
                    elif cours_s1[0][0] == "TD":
                        feuille.cell(row=ligne_insertion, column=7, value=(cours_s1[0][1] * 2 * liste_vac[i][4]))
                    elif cours_s1[0][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=8, value=(cours_s1[0][1] * 2 * liste_vac[i][5]))
                        feuille.cell(row=ligne_insertion, column=9, value=(cours_s1[0][1] * 2 * liste_vac[i][6]))

                if len(cours_s1) > 1:
                    if cours_s1[1][0] == "TD":
                        feuille.cell(row=ligne_insertion, column=7, value=(cours_s1[1][1] * 2 * liste_vac[i][4]))
                    elif cours_s1[1][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=8, value=(cours_s1[1][1] * 2 * liste_vac[i][5]))
                        feuille.cell(row=ligne_insertion, column=9, value=(cours_s1[1][1] * 2 * liste_vac[i][6]))

                if len(cours_s1) > 2:
                    if cours_s1[2][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=8, value=(cours_s1[2][1] * 2 * liste_vac[i][5]))
                        feuille.cell(row=ligne_insertion, column=9, value=(cours_s1[2][1] * 2 * liste_vac[i][6]))

                # On remplie les bonnes cases de janvier à aout

                if len(cours_s2) > 0:
                    if cours_s2[0][0] == "Cours":
                        feuille.cell(row=ligne_insertion, column=12, value=(cours_s2[0][1] * 2 * liste_vac[i][3]))
                    elif cours_s2[0][0] == "TD":
                        feuille.cell(row=ligne_insertion, column=13, value=(cours_s2[0][1] * 2 * liste_vac[i][4]))
                    elif cours_s2[0][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=14, value=(cours_s2[0][1] * 2 * liste_vac[i][5]))
                        feuille.cell(row=ligne_insertion, column=15, value=(cours_s2[0][1] * 2 * liste_vac[i][6]))

                if len(cours_s2) > 1:
                    if cours_s2[1][0] == "TD":
                        feuille.cell(row=ligne_insertion, column=13, value=(cours_s2[1][1] * 2 * liste_vac[i][4]))
                    elif cours_s2[1][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=14, value=(cours_s2[1][1] * 2 * liste_vac[i][5]))
                        feuille.cell(row=ligne_insertion, column=15, value=(cours_s2[1][1] * 2 * liste_vac[i][6]))

                if len(cours_s2) > 2:
                    if cours_s2[2][0] == "TP":
                        feuille.cell(row=ligne_insertion, column=14, value=(cours_s2[2][1] * 2 * liste_vac[i][5]))
                        feuille.cell(row=ligne_insertion, column=15, value=(cours_s2[2][1] * 2 * liste_vac[i][6]))

                valeur_cellule_6 = feuille.cell(row=ligne_insertion, column=6).value
                valeur_cellule_7 = feuille.cell(row=ligne_insertion, column=7).value
                valeur_cellule_8 = feuille.cell(row=ligne_insertion, column=8).value
                valeur_cellule_9 = feuille.cell(row=ligne_insertion, column=9).value

                # Vérification si les cellules ne sont pas vides
                if valeur_cellule_6 is None:
                    valeur_cellule_6 = 0
                if valeur_cellule_7 is None:
                    valeur_cellule_7 = 0
                if valeur_cellule_8 is  None :
                    valeur_cellule_8 = 0
                if valeur_cellule_9 is None:
                    valeur_cellule_9 = 0

                # Calcul de hetd_ares_1
                hetd_ares_1 = valeur_cellule_8 + valeur_cellule_9

                # Assignation de la valeur à la cellule 10
                feuille.cell(row=ligne_insertion, column=10, value=hetd_ares_1)

                hetd_total_1 = valeur_cellule_6 * 1.5 + valeur_cellule_7 + valeur_cellule_8 + valeur_cellule_9
                feuille.cell(row=ligne_insertion, column=11, value=hetd_total_1)

                valeur_cellule_12 = feuille.cell(row=ligne_insertion, column=12).value
                valeur_cellule_13 = feuille.cell(row=ligne_insertion, column=13).value
                valeur_cellule_14 = feuille.cell(row=ligne_insertion, column=14).value
                valeur_cellule_15 = feuille.cell(row=ligne_insertion, column=15).value

                if valeur_cellule_12 is None:
                    valeur_cellule_12 = 0
                if valeur_cellule_13 is None:
                    valeur_cellule_13 = 0
                if valeur_cellule_14 is None:
                    valeur_cellule_14 = 0
                if valeur_cellule_15 is None:
                    valeur_cellule_15 = 0


                # Calcul de hetd_ares_1
                hetd_ares_2 = valeur_cellule_14 + valeur_cellule_15

                # Assignation de la valeur à la cellule 16
                feuille.cell(row=ligne_insertion, column=16, value=hetd_ares_2)

                hetd_total_1 = valeur_cellule_12 * 1.5 + valeur_cellule_13 + valeur_cellule_14 + valeur_cellule_15
                feuille.cell(row=ligne_insertion, column=17, value=hetd_total_1)

                valeur_cellule_11 = feuille.cell(row=ligne_insertion, column=11).value
                valeur_cellule_17 = feuille.cell(row=ligne_insertion, column=17).value
                if valeur_cellule_11 is None:
                    valeur_cellule_11 = 0
                if valeur_cellule_17 is None:
                    valeur_cellule_17 = 0

                hetd_final = valeur_cellule_11 + valeur_cellule_17
                feuille.cell(row=ligne_insertion, column=18, value=hetd_final)


        # fonction qui trouve la première ligne dans laquelle se trouve une cellule possédant la donnée recherchée
        def trouver_ligne(feuille, contenu_cible):
            for row_number, row in enumerate(feuille.iter_rows(values_only=True), start=1):
                if contenu_cible in row:
                    return row_number

            return None

        # recuperation des données pour chaque ressource existante
        requeteRemplissage = "SELECT Ressource, CM, TD, TP, Acronyme FROM PLANRESSOURCE WHERE Ressource NOT LIKE '%SAE%' GROUP BY Ressource ORDER BY Ressource "
        cursor.execute(requeteRemplissage)
        liste_res = cursor.fetchall()

        for i in range(len(liste_res)):
            Remplissage(liste_res[i][0], liste_res[i][1], liste_res[i][2], liste_res[i][3], liste_res[i][4])

        # Enregistrer le classeur modifié dans un nouveau fichier
        classeur.save("./Excels ressources/FichierRessources.xlsx")


def main():
    # Instanciation de la classe GenerationRessources
    ressource_generator = GenerationRessources()
    # Appel de la méthode run()
    ressource_generator.run()


if __name__ == "__main__":
    main()
