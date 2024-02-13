import openpyxl
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import Font
import sqlite3

# connection à la base de donnée
db_connection = sqlite3.connect("../SAE.db")
cursor = db_connection.cursor()

# Creer un nouveau classeur Excel
classeur = openpyxl.Workbook()

# Selectionner la feuille active
feuille = classeur.active

def Remplissage(res, cmtot, tdtot, tptot, resp):
    # Créer une nouvelle feuille pour chaque matière
    feuille = classeur.create_sheet(res)

    # Ajouter des donnees a la feuille
    feuille['A1'] = 'Ressource'
    feuille['A1'].font = Font(bold=True)

    feuille['F1'] = 'Responsable'
    feuille['F1'].font = Font(bold=True)

    feuille['A3'] = 'Maquette'
    feuille['B3'] = 'CM (h)'
    feuille['C3'] = 'TD (h)'
    feuille['D3'] = 'TP (h)'

    feuille['A6'] = 'Intervenants'
    feuille['A6'].font = Font(bold=True)

    feuille['A8'] = 'Repartition'
    feuille['A8'].font = Font(bold=True)
    feuille['B8'] = 'Nombre de Groupes'
    feuille['A9'] = 'CM'
    feuille['A11'] = 'TD'
    feuille['A13'] = 'TP dedoubles'
    feuille['A15'] = 'TP non dedoubles'

    feuille['A17'] = 'Organisation detaillee'
    feuille['A17'].font = Font(bold=True)
    feuille['B19'] = 'CM heures'
    feuille['C19'] = 'TD heures'
    feuille['D19'] = 'TP heures'
    feuille['E19'] = 'Test'

    feuille['A21'] = 'Service previsionnel vacataires'
    feuille['A22'] = 'Nom'
    feuille['C22'] = 'BUT1/BUT2/BUT3'
    feuille['D22'] = 'ParcoursA  ParcoursB'
    feuille['E22'] = 'FI  FA'
    feuille['F22'] = 'Nombres d\'heures prevuesde septembre a decembre'
    feuille['L22'] = 'Nombres d\'heures prevuesde janvier a aout'
    feuille['F23'] = 'CM'
    feuille['G23'] = 'TD'
    feuille['H23'] = 'TP(en 1/2 groupe)'
    feuille['I23'] = 'TP(nondedouble)'
    feuille['J23'] = 'TP adeclarerARES'
    feuille['K23'] = 'Total enHETD'
    feuille['L23'] = 'CM'
    feuille['M23'] = 'TD'
    feuille['N23'] = 'TP(en 1/2 groupe)'
    feuille['O23'] = 'TP(nondedouble)'
    feuille['P23'] = 'TP adeclarerARES'
    feuille['Q23'] = 'Total enHETD'
    feuille['R23'] = 'Total enHETD'

    feuille.merge_cells('A22:B23')
    feuille.merge_cells('C22:C23')
    feuille.merge_cells('D22:D23')
    feuille.merge_cells('E22:E23')
    feuille.merge_cells('F22:K22')
    feuille.merge_cells('L22:Q22')

    feuille['A25'] = 'Service previsionnel titulaires'
    feuille['A26'] = 'Nom'
    feuille['C26'] = 'BUT1/BUT2/BUT3'
    feuille['D26'] = 'ParcoursA  ParcoursB'
    feuille['E26'] = 'FI  FA'
    feuille['F26'] = "Nombres d'heures prevues de septembre a decembre"
    feuille['L26'] = "Nombres d'heures prevues de janvier a aout"
    feuille['F27'] = 'CM'
    feuille['G27'] = 'TD'
    feuille['H27'] = 'TP(en 1/2 groupe)'
    feuille['I27'] = 'TP(nondedouble)'
    feuille['J27'] = 'TP a declarer ARES'
    feuille['K27'] = 'Total en HETD'
    feuille['L27'] = 'CM'
    feuille['M27'] = 'TD'
    feuille['N27'] = 'TP(en 1/2 groupe)'
    feuille['O27'] = 'TP(nondedouble)'
    feuille['P27'] = 'TP a declarer ARES'
    feuille['Q27'] = 'Total en HETD'
    feuille['R27'] = 'Total en HETD'

    feuille.merge_cells('A26:B27')
    feuille.merge_cells('C26:C27')
    feuille.merge_cells('D26:D27')
    feuille.merge_cells('E26:E27')
    feuille.merge_cells('F26:K26')
    feuille.merge_cells('L26:Q26')


    #nom responsable de ressource
    cellule_selectionnee = feuille['H1']
    cellule_selectionnee.value = resp

    #nom ressource
    cellule_selectionnee = feuille['C1']
    cellule_selectionnee.value = res

    #volume horaire total
    cellule_selectionnee = feuille['B4']
    cellule_selectionnee.value = cmtot

    cellule_selectionnee = feuille['C4']
    cellule_selectionnee.value = tdtot

    cellule_selectionnee = feuille['D4']
    cellule_selectionnee.value = tptot

    #Récupération des donnees des intervenants
    requete = "SELECT Intervenant, Acronyme, CM, NombreGroupes FROM DONNEEPROF WHERE SUBSTR(MatiereActuelle, 1, INSTR(MatiereActuelle, ' ') - 1) = ?"
    cursor.execute(requete, (res,))
    liste_inter = cursor.fetchall()


    # insertion données dans la feuille
    for i in range(len(liste_inter)):

        #nom et acronyme
        inter = trouver_ligne(feuille, "Intervenants")
        
        ligne_insertion = 7 + i
        feuille.insert_rows(ligne_insertion)
        feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][0])
        feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][1])

        # pourcentage du total de cm effectué par le professeur
        cms = trouver_ligne(feuille, "CM")

        ligne_insertion = cms + 1 + i
        feuille.insert_rows(ligne_insertion)
        feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][1])
        feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][2])

        tds = trouver_ligne(feuille, "TD")

        #nombre de groupes en TD
        ligne_insertion = tds + 1 + i
        feuille.insert_rows(ligne_insertion)
        feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][1])
        feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][3])

        tpsD = trouver_ligne(feuille, "TP dedoubles")

        #nombre de groupes en TP dédoublés
        ligne_insertion = tpsD + 1 + i
        feuille.insert_rows(ligne_insertion)
        feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][1])
        feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][3])

        tps = trouver_ligne(feuille, "TP non dedoubles")

        #nombre de groupes en TP non dédoublés
        ligne_insertion = tps + 1 + i
        feuille.insert_rows(ligne_insertion)
        feuille.cell(row=ligne_insertion, column=1, value=liste_inter[i][1])
        feuille.cell(row=ligne_insertion, column=2, value=liste_inter[i][3])


    #Récupération des données de cours par semaine 
    cours = "SELECT Semaine, TypeCours, COUNT(*) AS NombreCours FROM PLANINFO WHERE Ressource = ? GROUP BY Semaine, TypeCours "
    cursor.execute(cours, (res,))
    liste_cours = cursor.fetchall()


    for i in range(len(liste_cours)):

        hcm = trouver_ligne(feuille, "CM heures")

        ligne_insertion = hcm + 1 + i
        feuille.insert_rows(ligne_insertion)
        feuille.cell(row=ligne_insertion, column=1, value=liste_cours[i][0])

        ligne_date = trouver_ligne(feuille, liste_cours[i][0])
        if liste_cours[i][1] == "Cours" :
            feuille.cell(row=ligne_date, column=2, value=(liste_cours[i][2]) * 2)
        elif liste_cours[i][1] == "TD":
            feuille.cell(row=ligne_date, column=3, value=(liste_cours[i][2] * 2))
        elif liste_cours[i][1] == "TP":
            feuille.cell(row=ligne_date, column=4, value=(liste_cours[i][2] * 2))
        else:
            feuille.cell(row=ligne_date, column=5, value=(liste_cours[i][2] * 2))




    #Récupération des données pour remplir tableau titulaire
    titu = "SELECT Intervenant, Feuille_title, SUBSTR(MatiereActuelle, 1, INSTR(MatiereActuelle, ' ') - 1) AS matiere,d.CM, d.TD, TPD, TDNonD, Test, p.CM FROM DONNEEPROF d JOIN PLANRESSOURCE p ON matiere = Ressource WHERE AlerteProf == 1 AND matiere = ?"
    cursor.execute(titu, (res,))
    liste_titu = cursor.fetchall()

    ligne_titu = trouver_ligne(feuille, "Service previsionnel titulaires")

    for i in range(len(liste_titu)):
        ligne_insertion = ligne_titu + i + 3
        feuille.insert_rows(ligne_insertion)
        feuille.cell(row=ligne_insertion, column=1, value=(liste_titu[i][0]))
        if "S1" or "S2" in liste_titu[i][1]:
            feuille.cell(row=ligne_insertion, column=3, value="BUT1")
        elif "S3" or "S4" in liste_titu[i][1]:
            feuille.cell(row=ligne_insertion, column=3, value="BUT2")
        else:
            feuille.cell(row=ligne_insertion, column=3, value="BUT3")
        if "A" in liste_titu[i][1]:
            feuille.cell(row=ligne_insertion, column=4, value="A")
        elif "B" in liste_titu[i][1]:
            feuille.cell(row=ligne_insertion, column=4, value="B")
        if "S1" in liste_titu[i][1] or "S3" in liste_titu[i][1] or "S6" in liste_titu[i][1]:
            feuille.cell(row=ligne_insertion, column=6, value=(liste_titu[i][3] * liste_titu[i][8]) * 2)
            feuille.cell(row=ligne_insertion, column=7, value=liste_titu[i][4] * 2)
            feuille.cell(row=ligne_insertion, column=8, value=liste_titu[i][5] * 2)
            feuille.cell(row=ligne_insertion, column=9, value=liste_titu[i][6] * 2)
        else:
            feuille.cell(row=ligne_insertion, column=12, value=(liste_titu[i][3] * liste_titu[i][8]) * 2)
            feuille.cell(row=ligne_insertion, column=13, value=liste_titu[i][4] * 2)
            feuille.cell(row=ligne_insertion, column=14, value=liste_titu[i][5] * 2)
            feuille.cell(row=ligne_insertion, column=15, value=liste_titu[i][6] * 2)

        valeur_cellule_8 = feuille.cell(row=ligne_insertion, column=8).value
        valeur_cellule_9 = feuille.cell(row=ligne_insertion, column=9).value

        # Vérification si les cellules ne sont pas vides
        if valeur_cellule_8 is not None and valeur_cellule_9 is not None:
            # Calcul de hetd_ares_1
            hetd_ares_1 = valeur_cellule_8 + valeur_cellule_9

            # Assignation de la valeur à la cellule 10
            feuille.cell(row=ligne_insertion, column=10, value=hetd_ares_1)

        valeur_cellule_6 = feuille.cell(row=ligne_insertion, column=6).value
        valeur_cellule_7 = feuille.cell(row=ligne_insertion, column=7).value
        # Vérification si les cellules ne sont pas vides
        if valeur_cellule_8 is not None and valeur_cellule_9 is not None:
            hetd_total_1 = valeur_cellule_6 * 1.5 + valeur_cellule_7 + valeur_cellule_8 + valeur_cellule_9
            feuille.cell(row=ligne_insertion, column=11, value=hetd_total_1)

        valeur_cellule_11 =feuille.cell(row=ligne_insertion, column=11).value
        valeur_cellule_17 = feuille.cell(row=ligne_insertion, column=17).value

        if valeur_cellule_11 is not None and valeur_cellule_17 is not None:
            hetd_final = valeur_cellule_11 + valeur_cellule_17
        else:
            hetd_final = valeur_cellule_11 if valeur_cellule_11 is not None else valeur_cellule_17
            feuille.cell(row=ligne_insertion, column=18, value=hetd_final)


    # Récupération des données pour remplir tableau vacataire
    vac = "SELECT Intervenant, Feuille_title, SUBSTR(MatiereActuelle, 1, INSTR(MatiereActuelle, ' ') - 1) AS matiere,d.CM, d.TD, TDNonD, TPD, Test, p.CM FROM DONNEEPROF d JOIN PLANRESSOURCE p ON matiere = Ressource WHERE AlerteProf == 0 AND matiere = ?"
    cursor.execute(vac, (res,))
    liste_vac = cursor.fetchall()

    ligne_vac = trouver_ligne(feuille, "Service previsionnel vacataires")

    for i in range(len(liste_vac)):
        ligne_insertion = ligne_vac + i + 3
        feuille.insert_rows(ligne_insertion)
        feuille.cell(row=ligne_insertion, column=1, value=(liste_vac[i][0]))
        if "S1" in liste_vac[i][1] or "S2" in liste_vac[i][1]:
            feuille.cell(row=ligne_insertion, column=3, value="BUT1")
        elif "S3" in liste_vac[i][1] or "S4" in liste_vac[i][1]:
            feuille.cell(row=ligne_insertion, column=3, value="BUT2")
        else:
            feuille.cell(row=ligne_insertion, column=3, value="BUT3")
        if "A" in liste_vac[i][1]:
            feuille.cell(row=ligne_insertion, column=4, value="A")
        elif "B" in liste_vac[i][1]:
            feuille.cell(row=ligne_insertion, column=4, value="B")

        if "S1" in liste_vac[i][1] or "S3" in liste_vac[i][1] or "S6" in liste_vac[i][1]:
            feuille.cell(row=ligne_insertion, column=6, value=(liste_vac[i][3] * liste_vac[i][8]) * 2)
            feuille.cell(row=ligne_insertion, column=7, value=liste_vac[i][4] * 2)
            feuille.cell(row=ligne_insertion, column=8, value=liste_vac[i][5] * 2)
            feuille.cell(row=ligne_insertion, column=9, value=liste_vac[i][6] * 2)

        else:
            feuille.cell(row=ligne_insertion, column=12, value=(liste_vac[i][3] * liste_vac[i][8]) * 2)
            feuille.cell(row=ligne_insertion, column=13, value=liste_vac[i][4] * 2)
            feuille.cell(row=ligne_insertion, column=14, value=liste_vac[i][5] * 2)
            feuille.cell(row=ligne_insertion, column=15, value=liste_vac[i][6] * 2)


        valeur_cellule_8 = feuille.cell(row=ligne_insertion, column=8).value
        valeur_cellule_9 = feuille.cell(row=ligne_insertion, column=9).value

        # Vérification si les cellules ne sont pas vides
        if valeur_cellule_8 is not None and valeur_cellule_9 is not None:
            # Calcul de hetd_ares_1
            hetd_ares_1 = valeur_cellule_8 + valeur_cellule_9 * 2 / 3

            # Assignation de la valeur à la cellule 10
            feuille.cell(row=ligne_insertion, column=10, value=hetd_ares_1)

        valeur_cellule_6 = feuille.cell(row=ligne_insertion, column=6).value
        valeur_cellule_7 = feuille.cell(row=ligne_insertion, column=7).value
        # Vérification si les cellules ne sont pas vides
        if valeur_cellule_8 is not None and valeur_cellule_9 is not None:
            hetd_total = valeur_cellule_6 * 1.5 + valeur_cellule_7 + valeur_cellule_8 * 2/3 + valeur_cellule_9
            feuille.cell(row=ligne_insertion, column=11, value=hetd_total)

        valeur_cellule_11 = feuille.cell(row=ligne_insertion, column=11).value
        valeur_cellule_17 = feuille.cell(row=ligne_insertion, column=17).value

        if valeur_cellule_11 is not None and valeur_cellule_17 is not None:
            hetd_final = valeur_cellule_11 + valeur_cellule_17
        else:
            hetd_final = valeur_cellule_11 if valeur_cellule_11 is not None else valeur_cellule_17
            feuille.cell(row=ligne_insertion, column=18, value=hetd_final)





#fonction qui trouve la première ligne dans laquelle se trouve une cellule possédant la donnée recherchée
def trouver_ligne(feuille, contenu_cible):
    for row_number, row in enumerate(feuille.iter_rows(values_only=True), start=1):
        if contenu_cible in row:
            return row_number
    return None

#recuperation des données pour chaque ressource existante
requeteRemplissage = "SELECT Ressource, CM, TD, TP, Acronyme FROM PLANRESSOURCE WHERE Ressource NOT LIKE '%SAE%' GROUP BY Ressource ORDER BY Ressource "
cursor.execute(requeteRemplissage)
liste_res = cursor.fetchall()

for i in range(len(liste_res)):
     Remplissage(liste_res[i][0], liste_res[i][1], liste_res[i][2], liste_res[i][3], liste_res[i][4])


# Enregistrer le classeur modifié dans un nouveau fichier
classeur.save('Ressources.xlsx')
